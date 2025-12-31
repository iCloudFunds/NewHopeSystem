from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import User, Group
from django.utils.crypto import get_random_string
from django.contrib import admin
from django import forms
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib import messages
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from io import BytesIO
from django.contrib.auth.hashers import make_password
import datetime

from .models import Student, Department, Class, Subject, Teacher

class ImportExcelForm(forms.Form):
    excel_file = forms.FileField(
        label='Select Excel File',
        help_text='File must be .xlsx format with columns matching the required schema.'
    )

@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):
    # Fields to display in list view - REMOVED 'status' from list_display
    list_display = ('student_id', 'full_name', 'current_class', 'date_of_admission')
    search_fields = ('student_id', 'full_name', 'parent_guardian_name')
    list_filter = ('current_class',)  # REMOVED 'status' from list_filter
    
    # Fields organization in add/edit form
    fieldsets = (
        ('STUDENT INFORMATION', {
            'fields': ('full_name', 'date_of_birth', 'date_of_admission', 'current_class', 'address')
        }),
        ('PARENT/GUARDIAN INFORMATION', {
            'fields': ('parent_guardian_name', 'parent_guardian_contact'),
            'classes': ('collapse',)  # Makes this section collapsible
        }),
        ('SYSTEM INFORMATION (Auto-filled)', {
            'fields': ('user', 'student_id'),  # REMOVED 'status' from fields
            'classes': ('collapse',),
            'description': 'These fields are managed by the system'
        }),
    )
    
    # Make user field read-only (admin can see but not change)
    readonly_fields = ('student_id',)
    
    # Updated actions list with all new actions - REMOVED 'bulk_update_student_status'
    actions = [
        'bulk_send_whatsapp_notifications',
        'generate_individual_student_profile_reports',
        'export_student_data_to_excel',
        'bulk_update_student_contact_information',
        'generate_student_id_cards',
        'assign_students_to_multiple_classes',
        'create_student_groups',
        'generate_parent_meeting_schedules',
        'import_export_student_photos',  # REMOVED: 'bulk_update_student_status',
        'generate_fee_payment_report',
        'generate_academic_performance_report', 
        'generate_discipline_report',
        'promote_to_next_class',
        'send_whatsapp_report',
        'export_as_pdf',
        'delete_selected'
    ]
    
    # Keep your existing change_list_template
    change_list_template = 'admin/core/student/student_change_list.html'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.import_excel, name='import_excel'),
            path('import-photos/', self.import_student_photos, name='import_student_photos'),
        ]
        return custom_urls + urls
    
    # 1. Bulk Send WhatsApp Notifications to Parents/Guardians
    def bulk_send_whatsapp_notifications(self, request, queryset):
        if request.method == 'POST':
            message = request.POST.get('message', '')
            sent_count = 0
            failed_count = 0
            
            for student in queryset:
                if student.parent_guardian_contact:
                    try:
                        # WhatsApp integration placeholder
                        # In production, you would use WhatsApp Business API
                        phone_number = self._extract_phone_number(student.parent_guardian_contact)
                        
                        # Log the message (for now, just print to console)
                        print(f"WhatsApp to {student.parent_guardian_name}: {message}")
                        sent_count += 1
                    except Exception as e:
                        print(f"Failed to send to {student.full_name}: {str(e)}")
                        failed_count += 1
            
            messages.success(
                request,
                f"Successfully sent WhatsApp notifications to {sent_count} parents. "
                f"Failed: {failed_count}. Note: Actual WhatsApp integration requires API setup."
            )
            return redirect('../')
        
        # Show form to compose message
        context = {
            'students': queryset,
            'title': 'Send WhatsApp Notifications to Parents/Guardians'
        }
        return render(request, 'admin/send_whatsapp_form.html', context)
    
    bulk_send_whatsapp_notifications.short_description = "📱 Bulk send WhatsApp notifications to parents"
    
    # Helper method to extract phone number
    def _extract_phone_number(self, contact_string):
        # Remove non-numeric characters
        import re
        numbers = re.findall(r'\d+', contact_string)
        if numbers:
            # Take the longest number found
            return max(numbers, key=len)
        return None
    
    # 2. Generate Individual Student Profile Reports
    def generate_individual_student_profile_reports(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        for student in queryset:
            # Start new page for each student
            p.setFont("Helvetica-Bold", 16)
            p.drawString(100, 750, f"Student Profile Report")
            
            # Student ID and Basic Info
            p.setFont("Helvetica", 12)
            y_position = 700
            p.drawString(100, y_position, f"Student ID: {student.student_id}")
            y_position -= 20
            p.drawString(100, y_position, f"Full Name: {student.full_name}")
            y_position -= 20
            p.drawString(100, y_position, f"Date of Birth: {student.date_of_birth}")
            y_position -= 20
            p.drawString(100, y_position, f"Admission Date: {student.date_of_admission}")
            y_position -= 20
            p.drawString(100, y_position, f"Current Class: {student.current_class}")
            
            # Parent/Guardian Information
            y_position -= 40
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_position, "Parent/Guardian Information:")
            y_position -= 20
            p.setFont("Helvetica", 12)
            p.drawString(100, y_position, f"Name: {student.parent_guardian_name}")
            y_position -= 20
            p.drawString(100, y_position, f"Contact: {student.parent_guardian_contact}")
            
            # Address
            y_position -= 40
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_position, "Address:")
            y_position -= 20
            p.setFont("Helvetica", 12)
            # Wrap address text
            address_lines = self._wrap_text(student.address, 60)
            for line in address_lines:
                if y_position < 50:
                    p.showPage()
                    p.setFont("Helvetica", 12)
                    y_position = 750
                p.drawString(100, y_position, line)
                y_position -= 15
            
            # Add a separator between students
            if queryset.count() > 1 and student != queryset.last():
                p.showPage()
                p.setFont("Helvetica", 12)
        
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="student_profiles.pdf"'
        return response
    
    generate_individual_student_profile_reports.short_description = "📄 Generate individual student profile reports"
    
    # Helper method for text wrapping
    def _wrap_text(self, text, max_length):
        if not text:
            return ['']
        words = text.split()
        lines = []
        current_line = []
        current_length = 0
        
        for word in words:
            if current_length + len(word) + 1 <= max_length:
                current_line.append(word)
                current_length += len(word) + 1
            else:
                lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word)
        
        if current_line:
            lines.append(' '.join(current_line))
        
        return lines
    
    # 3. Export Student Data to Excel
    def export_student_data_to_excel(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students Data"
        
        # Headers - REMOVED 'Status' column
        headers = [
            'Student ID', 'Full Name', 'Date of Birth', 'Date of Admission',
            'Current Class', 'Parent/Guardian Name', 'Parent/Guardian Contact',
            'Address', 'Email', 'Username'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws.column_dimensions[col_letter].width = 20
        
        # Data rows - REMOVED status field
        for row_num, student in enumerate(queryset, 2):
            ws[f'A{row_num}'] = student.student_id
            ws[f'B{row_num}'] = student.full_name
            ws[f'C{row_num}'] = str(student.date_of_birth) if student.date_of_birth else ''
            ws[f'D{row_num}'] = str(student.date_of_admission) if student.date_of_admission else ''
            ws[f'E{row_num}'] = str(student.current_class) if student.current_class else ''
            ws[f'F{row_num}'] = student.parent_guardian_name
            ws[f'G{row_num}'] = student.parent_guardian_contact
            ws[f'H{row_num}'] = student.address
            ws[f'I{row_num}'] = student.user.email if student.user else ''
            ws[f'J{row_num}'] = student.user.username if student.user else ''
        
        # Save to response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
        return response
    
    export_student_data_to_excel.short_description = "📊 Export student data to Excel"
    
    # 4. Bulk Update Student Contact Information
    def bulk_update_student_contact_information(self, request, queryset):
        if request.method == 'POST':
            field_name = request.POST.get('field_name')
            new_value = request.POST.get('new_value')
            updated_count = 0
            
            for student in queryset:
                try:
                    if field_name == 'parent_guardian_name':
                        student.parent_guardian_name = new_value
                    elif field_name == 'parent_guardian_contact':
                        student.parent_guardian_contact = new_value
                    elif field_name == 'address':
                        student.address = new_value
                    student.save()
                    updated_count += 1
                except Exception as e:
                    messages.error(request, f"Error updating {student.full_name}: {str(e)}")
            
            messages.success(request, f"Successfully updated contact information for {updated_count} students.")
            return redirect('../')
        
        # Show form
        context = {
            'students': queryset,
            'title': 'Bulk Update Student Contact Information'
        }
        return render(request, 'admin/bulk_update_contact_form.html', context)
    
    bulk_update_student_contact_information.short_description = "📝 Bulk update student contact information"
    
    # 5. Generate Student ID Cards
    def generate_student_id_cards(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=(300, 200))  # ID card size
        
        for student in queryset:
            # ID Card Border
            p.setStrokeColorRGB(0, 0, 0)  # Black border
            p.rect(10, 10, 280, 180)
            
            # School Name
            p.setFont("Helvetica-Bold", 14)
            p.drawString(50, 160, "New Hope College")
            
            # Student Photo Placeholder
            p.setStrokeColorRGB(0.8, 0.8, 0.8)  # Light gray
            p.rect(20, 70, 80, 100)  # Photo area
            p.setFont("Helvetica", 10)
            p.drawString(30, 125, "PHOTO")
            
            # Student Information
            p.setFont("Helvetica", 10)
            p.drawString(120, 140, f"ID: {student.student_id}")
            p.drawString(120, 120, f"Name: {student.full_name}")
            p.drawString(120, 100, f"Class: {student.current_class}")
            p.drawString(120, 80, f"Valid: {datetime.date.today().year}")
            
            # Barcode placeholder
            p.setStrokeColorRGB(0.8, 0.8, 0.8)
            p.rect(120, 30, 150, 40)
            p.setFont("Helvetica", 8)
            p.drawString(170, 50, f"BARCODE: {student.student_id}")
            
            # Footer
            p.setFont("Helvetica-Oblique", 8)
            p.drawString(20, 20, "Issued by New Hope College Administration")
            
            p.showPage()  # New page for next ID card
        
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="student_id_cards.pdf"'
        return response
    
    generate_student_id_cards.short_description = "🪪 Generate student ID cards"
    
    # 6. Assign Students to Multiple Classes
    def assign_students_to_multiple_classes(self, request, queryset):
        if request.method == 'POST':
            class_id = request.POST.get('class_id')
            try:
                target_class = Class.objects.get(id=class_id)
                updated_count = 0
                
                for student in queryset:
                    student.current_class = target_class
                    student.save()
                    updated_count += 1
                
                messages.success(request, f"Successfully assigned {updated_count} students to {target_class.name}.")
                return redirect('../')
            except Class.DoesNotExist:
                messages.error(request, "Selected class does not exist.")
        
        # Show form with class selection
        classes = Class.objects.all()
        context = {
            'students': queryset,
            'classes': classes,
            'title': 'Assign Students to Class'
        }
        return render(request, 'admin/assign_students_to_class_form.html', context)
    
    assign_students_to_multiple_classes.short_description = "📚 Assign students to classes"
    
    # 7. Create Student Groups/Tagging System
    def create_student_groups(self, request, queryset):
        if request.method == 'POST':
            group_name = request.POST.get('group_name', '').strip()
            group_description = request.POST.get('group_description', '')
            
            if not group_name:
                messages.error(request, "Group name is required.")
            else:
                # Note: You would need a StudentGroup model for this
                # For now, we'll log it
                student_ids = [str(student.id) for student in queryset]
                print(f"Created group '{group_name}' with students: {', '.join(student_ids)}")
                messages.success(
                    request,
                    f"Group '{group_name}' created with {len(queryset)} students. "
                    "Note: This feature requires StudentGroup model implementation."
                )
                return redirect('../')
        
        context = {
            'students': queryset,
            'title': 'Create Student Group'
        }
        return render(request, 'admin/create_student_group_form.html', context)
    
    create_student_groups.short_description = "🏷️ Create student groups"
    
    # 8. Generate Parent Meeting Schedules
    def generate_parent_meeting_schedules(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 750, "Parent-Teacher Meeting Schedule")
        p.setFont("Helvetica", 10)
        p.drawString(400, 750, f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d')}")
        
        y_position = 700
        time_slot = datetime.time(9, 0)  # Start at 9:00 AM
        
        for student in queryset:
            if y_position < 100:
                p.showPage()
                p.setFont("Helvetica", 12)
                y_position = 750
            
            p.setFont("Helvetica-Bold", 12)
            p.drawString(100, y_position, f"{time_slot.strftime('%I:%M %p')} - {student.full_name}")
            y_position -= 15
            p.setFont("Helvetica", 11)
            p.drawString(100, y_position, f"Student ID: {student.student_id}")
            y_position -= 15
            p.drawString(100, y_position, f"Class: {student.current_class}")
            y_position -= 15
            p.drawString(100, y_position, f"Parent: {student.parent_guardian_name}")
            y_position -= 15
            p.drawString(100, y_position, f"Contact: {student.parent_guardian_contact}")
            y_position -= 30
            
            # Increment time slot by 15 minutes
            time_slot = (datetime.datetime.combine(datetime.date.today(), time_slot) + 
                        datetime.timedelta(minutes=15)).time()
        
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="parent_meeting_schedule.pdf"'
        return response
    
    generate_parent_meeting_schedules.short_description = "📅 Generate parent meeting schedules"
    
    # 9. REMOVED: Bulk Update Student Status (since Student model doesn't have status field)
    # This method has been removed
    
    # 10. Import/Export Student Photos
    def import_export_student_photos(self, request, queryset):
        # This would typically redirect to a custom view for photo management
        messages.info(
            request,
            f"Photo management for {queryset.count()} students. "
            "Note: This feature requires Student model to have a 'photo' ImageField and proper storage setup."
        )
        
        # For now, we'll create a simple export of student data with photo references
        import zipfile
        from io import BytesIO
        
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, 'w') as zip_file:
            # Add student information file
            info_content = "Student ID,Full Name,Photo Filename\n"
            for student in queryset:
                # Assuming photo field exists - adjust field name as needed
                photo_filename = f"{student.student_id}.jpg"
                info_content += f"{student.student_id},{student.full_name},{photo_filename}\n"
            
            zip_file.writestr("students_photos_list.csv", info_content)
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="student_photos_template.zip"'
        return response
    
    import_export_student_photos.short_description = "🖼️ Import/export student photos"
    
    def import_student_photos(self, request):
        # Placeholder for photo import functionality
        if request.method == 'POST':
            messages.info(request, "Photo import functionality requires implementation with file handling.")
            return redirect('../')
        
        context = {
            'title': 'Import Student Photos'
        }
        return render(request, 'admin/import_photos_form.html', context)
    
    # UPDATED: Excel import method - REMOVED status field
    def import_excel(self, request):
        if request.method == 'POST':
            form = ImportExcelForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                try:
                    df = pd.read_excel(excel_file, engine='openpyxl')
                    
                    # Updated required columns for new field structure
                    required_columns = {'full_name', 'email'}
                    if not required_columns.issubset(df.columns):
                        missing = required_columns - set(df.columns)
                        messages.error(request, f'Missing required columns: {missing}')
                    else:
                        success_count = 0
                        for _, row in df.iterrows():
                            # Create admin user for the student
                            user, created = User.objects.get_or_create(
                                username=row['email'],
                                defaults={
                                    'email': row['email'],
                                    'first_name': row['full_name'].split()[0] if ' ' in row['full_name'] else row['full_name'],
                                    'last_name': row['full_name'].split()[-1] if ' ' in row['full_name'] else '',
                                    'is_staff': True  # Admin privilege
                                }
                            )
                            
                            # Create or update the Student with new fields - REMOVED status
                            student, created = Student.objects.update_or_create(
                                user=user,
                                defaults={
                                    'full_name': row['full_name'],
                                    'date_of_birth': row.get('date_of_birth'),
                                    'date_of_admission': row.get('date_of_admission', datetime.date.today()),
                                    'parent_guardian_name': row.get('parent_guardian_name', ''),
                                    'parent_guardian_contact': row.get('parent_guardian_contact', ''),
                                    'address': row.get('address', ''),
                                    'current_class': row.get('current_class', None)
                                }
                            )
                            success_count += 1
                        messages.success(request, f'Successfully imported/updated {success_count} students.')
                        return HttpResponseRedirect('../')
                except Exception as e:
                    messages.error(request, f'Error processing file: {str(e)}')
        else:
            form = ImportExcelForm()
        
        # Updated schema for Excel import - REMOVED status column
        schema = [
            {'column': 'full_name', 'required': True, 'description': 'Student full name'},
            {'column': 'email', 'required': True, 'description': 'Email (used for admin login)'},
            {'column': 'date_of_birth', 'required': False, 'description': 'YYYY-MM-DD format (cannot be today)'},
            {'column': 'date_of_admission', 'required': False, 'description': 'YYYY-MM-DD format (defaults to today)'},
            {'column': 'parent_guardian_name', 'required': False, 'description': 'Parent/Guardian full name'},
            {'column': 'parent_guardian_contact', 'required': False, 'description': 'Phone number'},
            {'column': 'address', 'required': False, 'description': 'Full physical address'},
            {'column': 'current_class', 'required': False, 'description': 'Class ID (e.g., Form 1A)'},
        ]
        
        context = {
            'form': form,
            'schema': schema,
            'title': 'Import Students from Excel'
        }
        return render(request, 'admin/import_excel.html', context)

    # PDF Export Method - REMOVED status field
    def export_as_pdf(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # Set up PDF title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 750, "New Hope College - Student Report")
        
        # Add generation date
        p.setFont("Helvetica", 10)
        p.drawString(400, 750, f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Draw a line
        p.line(100, 735, 500, 735)
        
        # Add student information
        y_position = 700
        p.setFont("Helvetica", 12)
        
        for student in queryset:
            if y_position < 100:  # New page if running out of space
                p.showPage()
                p.setFont("Helvetica", 12)
                y_position = 750
            
            p.drawString(100, y_position, f"Student ID: {student.student_id}")
            y_position -= 20
            p.drawString(100, y_position, f"Name: {student.full_name}")
            y_position -= 20
            p.drawString(100, y_position, f"Class: {student.current_class}")
            y_position -= 20
            p.drawString(100, y_position, f"Admission Date: {student.date_of_admission}")
            y_position -= 30  # Extra space between students
            
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="student_report.pdf"'
        return response
    
    export_as_pdf.short_description = "Export selected students as PDF"

    # NEW ACTION: Generate Fee Payment Report
    def generate_fee_payment_report(self, request, queryset):
        from django.http import HttpResponse
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from io import BytesIO
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 750, "New Hope College - Fee Payment Report")
        
        y_position = 700
        p.setFont("Helvetica", 12)
        
        for student in queryset:
            p.drawString(100, y_position, f"Student: {student.full_name} ({student.student_id})")
            y_position -= 20
            p.drawString(100, y_position, f"Class: {student.current_class}")
            y_position -= 20
            p.drawString(100, y_position, "Fee Status: [To be implemented with Fee model]")
            y_position -= 40
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="fee_report.pdf"'
        return response
    
    generate_fee_payment_report.short_description = "📊 Generate fee payment report"

    # NEW ACTION: Generate Academic Performance Results
    def generate_academic_performance_report(self, request, queryset):
        messages.success(request, f"Academic performance reports queued for {queryset.count()} student(s).")
        # Future implementation: Generate report cards with grades
    generate_academic_performance_report.short_description = "📈 Generate academic performance results"

    # NEW ACTION: Generate Discipline Results  
    def generate_discipline_report(self, request, queryset):
        messages.info(request, f"Discipline reports prepared for {queryset.count()} student(s).")
        # Future implementation: Show disciplinary records
    generate_discipline_report.short_description = "⚖️ Generate discipline results"

    # NEW ACTION: Promote to Next Class
    def promote_to_next_class(self, request, queryset):
        from django.db.models import Q
        
        promoted_count = 0
        for student in queryset:
            if student.current_class:
                current_class = student.current_class
                
                # Determine next class based on current class name
                # This is a basic implementation - you'll need to customize the logic
                if 'Form 1' in current_class.name:
                    next_class_name = current_class.name.replace('Form 1', 'Form 2')
                elif 'Form 2' in current_class.name:
                    next_class_name = current_class.name.replace('Form 2', 'Form 3')
                elif 'Form 3' in current_class.name:
                    next_class_name = current_class.name.replace('Form 3', 'Form 4')
                elif 'Form 4' in current_class.name:
                    next_class_name = current_class.name.replace('Form 4', 'Form 5')
                elif 'Form 5' in current_class.name:
                    next_class_name = current_class.name.replace('Form 5', 'LowerSixth')
                else:
                    next_class_name = None
                
                if next_class_name:
                    # Find the next class
                    next_class = Class.objects.filter(
                        Q(name=next_class_name) | Q(name__contains=next_class_name),
                        academic_year=current_class.academic_year
                    ).first()
                    
                    if next_class:
                        student.current_class = next_class
                        student.save()
                        promoted_count += 1
        
        if promoted_count > 0:
            messages.success(request, f"Successfully promoted {promoted_count} student(s) to next class.")
        else:
            messages.warning(request, "No students could be promoted. Check class naming patterns.")
    
    promote_to_next_class.short_description = "⬆️ Promote exceptionally to next class"

    # NEW ACTION: Send WhatsApp Report
    def send_whatsapp_report(self, request, queryset):
        # Note: This requires WhatsApp Business API setup
        # This is a placeholder for the implementation
        
        for student in queryset:
            # Extract parent/guardian contact
            contact_info = student.parent_guardian_contact
            
            # In a real implementation, you would:
            # 1. Parse phone number from contact_info
            # 2. Use WhatsApp Business API to send message
            # 3. Log the sent message
            
            # Placeholder logic
            phone_number = "extract_from_contact_info"
            message = f"Report for {student.full_name} ({student.student_id})"
        
        messages.info(request, 
            f"WhatsApp reports prepared for {queryset.count()} student(s). "
            "Note: WhatsApp integration requires API setup."
        )
    
    send_whatsapp_report.short_description = "💬 Send reports to parents/guardians WhatsApp"


@admin.register(Department)
class DepartmentAdmin(admin.ModelAdmin):
    list_display = ('name',)
    search_fields = ('name',)
    
    # Add these actions to the existing class
    actions = [
        'export_department_list',
        'assign_department_head', 
        'generate_department_report',
        'view_department_classes',
        'view_department_teachers',
        'archive_departments',
        'delete_selected',
    ]
    
    # 1. Export Department List
    def export_department_list(self, request, queryset):
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="departments_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Department Name', 'Classes Count', 'Teachers Count', 'Students Count'])
        
        for department in queryset:
            classes_count = Class.objects.filter(department=department).count()
            teachers_count = Teacher.objects.filter(department=department).count()
            
            # Count students in this department through classes
            department_classes = Class.objects.filter(department=department)
            students_count = Student.objects.filter(current_class__in=department_classes).count()
            
            writer.writerow([
                department.name,
                classes_count,
                teachers_count,
                students_count
            ])
        
        return response
    
    export_department_list.short_description = "📥 Export department list"
    
    # 2. Assign Department Head
    def assign_department_head(self, request, queryset):
        if request.method == 'POST':
            teacher_id = request.POST.get('teacher_id')
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                for department in queryset:
                    # In a real implementation, you would have a 'head' field in Department model
                    # For now, we'll use the department name in teacher's notes
                    teacher.notes = f"Head of {department.name}"
                    teacher.save()
                
                self.message_user(
                    request,
                    f"Successfully assigned {teacher.full_name} as head of {queryset.count()} department(s).",
                    messages.SUCCESS
                )
                return None
            except Teacher.DoesNotExist:
                self.message_user(request, "Teacher not found.", messages.ERROR)
        
        teachers = Teacher.objects.filter(status='ACTIVE')
        return render(
            request,
            'admin/assign_department_head_form.html',
            {
                'departments': queryset,
                'teachers': teachers,
                'title': 'Assign Department Head'
            }
        )
    
    assign_department_head.short_description = "👨‍💼 Assign department head"
    
    # 3. Generate Department Report
    def generate_department_report(self, request, queryset):
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from io import BytesIO
        from django.http import HttpResponse
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 750, "Department Report")
        
        y_position = 700
        
        for department in queryset:
            if y_position < 100:
                p.showPage()
                p.setFont("Helvetica", 12)
                y_position = 750
            
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_position, f"Department: {department.name}")
            y_position -= 20
            
            # Get statistics
            classes = Class.objects.filter(department=department)
            teachers = Teacher.objects.filter(department=department)
            department_classes = Class.objects.filter(department=department)
            students = Student.objects.filter(current_class__in=department_classes)
            
            p.setFont("Helvetica", 12)
            p.drawString(100, y_position, f"Classes: {classes.count()}")
            y_position -= 20
            p.drawString(100, y_position, f"Teachers: {teachers.count()}")
            y_position -= 20
            p.drawString(100, y_position, f"Students: {students.count()}")
            y_position -= 40
            
            # List classes
            if classes.exists():
                p.setFont("Helvetica-Bold", 12)
                p.drawString(100, y_position, "Classes in this department:")
                y_position -= 20
                p.setFont("Helvetica", 10)
                
                for class_obj in classes[:10]:  # Limit to 10 classes
                    if y_position < 50:
                        p.showPage()
                        p.setFont("Helvetica", 10)
                        y_position = 750
                    
                    p.drawString(120, y_position, f"• {class_obj.name} ({class_obj.level})")
                    y_position -= 15
                
                if classes.count() > 10:
                    p.drawString(120, y_position, f"... and {classes.count() - 10} more classes")
                    y_position -= 20
            
            y_position -= 30
        
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="department_report.pdf"'
        return response
    
    generate_department_report.short_description = "📄 Generate department report"
    
    # 4. View Department Classes
    def view_department_classes(self, request, queryset):
        if queryset.count() == 1:
            department = queryset.first()
            url = reverse('admin:core_class_changelist') + f'?department__id__exact={department.id}'
            return redirect(url)
        else:
            self.message_user(request, "Please select only ONE department to view its classes.", messages.WARNING)
    
    view_department_classes.short_description = "🏫 View department classes"
    
    # 5. View Department Teachers
    def view_department_teachers(self, request, queryset):
        if queryset.count() == 1:
            department = queryset.first()
            url = reverse('admin:core_teacher_changelist') + f'?department__id__exact={department.id}'
            return redirect(url)
        else:
            self.message_user(request, "Please select only ONE department to view its teachers.", messages.WARNING)
    
    view_department_teachers.short_description = "👨‍🏫 View department teachers"
    
    # 6. Archive Departments
    def archive_departments(self, request, queryset):
        # Note: This requires adding an 'is_active' or 'archived' field to the Department model
        # For now, we'll use a placeholder approach
        archived_count = 0
        
        for department in queryset:
            # In a real implementation, you would have:
            # department.is_active = False  # or department.archived = True
            # department.save()
            archived_count += 1
        
        self.message_user(
            request,
            f"{archived_count} department(s) marked for archiving. "
            "Note: Add 'is_active' field to Department model for proper archiving.",
            messages.INFO
        )
    
    archive_departments.short_description = "📁 Archive departments"


@admin.register(Class)
class ClassAdmin(admin.ModelAdmin):
    list_display = ('name', 'department', 'level', 'stream', 'academic_year')
    list_filter = ('department', 'level', 'stream', 'academic_year')
    search_fields = ('name', 'academic_year')
    filter_horizontal = ('teachers',)
    
    actions = [
        'view_students_in_class',
        'view_subjects_for_class',  # NEW ACTION ADDED HERE
        'assign_teacher_to_class',
        'generate_class_timetable',
        'view_class_attendance_report',
        'promote_entire_class',
        'generate_class_performance_report',
        'duplicate_class_for_next_year',
    ]
    
    def view_students_in_class(self, request, queryset):
        if queryset.count() == 1:
            class_obj = queryset.first()
            url = reverse('admin:core_student_changelist') + f'?current_class__id__exact={class_obj.id}'
            return redirect(url)
        else:
            self.message_user(request, "Please select only ONE class to view its students.", messages.WARNING)
    view_students_in_class.short_description = "👥 View students in class"
    
    # NEW ACTION: View Subjects for Class
    def view_subjects_for_class(self, request, queryset):
        if queryset.count() == 1:
            class_obj = queryset.first()
            # Get all subjects taught by teachers assigned to this class
            teachers = class_obj.teachers.all()
            subjects = Subject.objects.none()
            
            for teacher in teachers:
                subjects = subjects | teacher.subjects.all()
            
            # Remove duplicates and order by name
            subjects = subjects.distinct().order_by('name')
            
            context = {
                'class_obj': class_obj,
                'teachers': teachers,
                'subjects': subjects,
                'title': f'Subjects for {class_obj.name}',
            }
            return render(request, 'admin/view_subjects_for_class.html', context)
        else:
            self.message_user(request, "Please select only ONE class to view its subjects.", messages.WARNING)
    view_subjects_for_class.short_description = "📚 View subjects for class"
    
    def assign_teacher_to_class(self, request, queryset):
        if request.method == 'POST' and 'teacher' in request.POST:
            teacher_id = request.POST['teacher']
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                for class_obj in queryset:
                    class_obj.teachers.add(teacher)
                self.message_user(
                    request,
                    f"Successfully assigned {teacher.full_name} to {queryset.count()} class(es).",
                    messages.SUCCESS
                )
                return None  # Return to class list
            except Teacher.DoesNotExist:
                self.message_user(request, "Teacher not found.", messages.ERROR)
        
        # Show form to select teacher
        teachers = Teacher.objects.filter(status='ACTIVE')
        return render(
            request,
            'admin/assign_teacher_form.html',
            {
                'classes': queryset,
                'teachers': teachers,
                'title': 'Assign Teacher to Selected Classes'
            }
        )
    assign_teacher_to_class.short_description = "👨‍🏫 Assign teacher to class"
    
    def generate_class_timetable(self, request, queryset):
        self.message_user(
            request,
            f"Timetable generation queued for {queryset.count()} class(es). "
            "Feature requires Timetable model implementation.",
            messages.INFO
        )
    generate_class_timetable.short_description = "📅 Generate class timetable"
    
    def view_class_attendance_report(self, request, queryset):
        self.message_user(
            request,
            f"Attendance reports prepared for {queryset.count()} class(es). "
            "Feature requires Attendance model implementation.",
            messages.INFO
        )
    view_class_attendance_report.short_description = "📊 View class attendance report"
    
    def promote_entire_class(self, request, queryset):
        promoted_count = 0
        for class_obj in queryset:
            # Find students in this class
            students = Student.objects.filter(current_class=class_obj)
            for student in students:
                # Determine next class (simplified logic)
                if class_obj.level == 'ORDINARY LEVEL':
                    # Example: Form 1A → Form 2A
                    next_name = class_obj.name.replace('Form 1', 'Form 2').replace('Form 2', 'Form 3').replace('Form 3', 'Form 4').replace('Form 4', 'Form 5')
                    try:
                        next_class = Class.objects.get(
                            name=next_name,
                            academic_year=class_obj.academic_year
                        )
                        student.current_class = next_class
                        student.save()
                        promoted_count += 1
                    except Class.DoesNotExist:
                        continue
        
        if promoted_count > 0:
            self.message_user(
                request,
                f"Successfully promoted {promoted_count} student(s) from {queryset.count()} class(es).",
                messages.SUCCESS
            )
        else:
            self.message_user(
                request,
                "No students could be promoted. Check class naming patterns.",
                messages.WARNING
            )
    promote_entire_class.short_description = "⬆️ Promote entire class"
    
    def generate_class_performance_report(self, request, queryset):
        self.message_user(
            request,
            f"Performance reports queued for {queryset.count()} class(es). "
            "Feature requires Grade/Result model implementation.",
            messages.INFO
        )
    generate_class_performance_report.short_description = "📈 Generate class performance report"
    
    def duplicate_class_for_next_year(self, request, queryset):
        created_count = 0
        for class_obj in queryset:
            # Calculate next academic year (e.g., 2024-2025 → 2025-2026)
            if '-' in class_obj.academic_year:
                start_year, end_year = map(int, class_obj.academic_year.split('-'))
                next_academic_year = f"{start_year+1}-{end_year+1}"
                
                # Check if class already exists for next year
                if not Class.objects.filter(
                    name=class_obj.name,
                    department=class_obj.department,
                    level=class_obj.level,
                    stream=class_obj.stream,
                    academic_year=next_academic_year
                ).exists():
                    
                    # Create new class for next year
                    new_class = Class.objects.create(
                        name=class_obj.name,
                        department=class_obj.department,
                        level=class_obj.level,
                        stream=class_obj.stream,
                        academic_year=next_academic_year
                    )
                    # Copy teachers
                    new_class.teachers.set(class_obj.teachers.all())
                    created_count += 1
        
        if created_count > 0:
            self.message_user(
                request,
                f"Successfully created {created_count} class(es) for next academic year.",
                messages.SUCCESS
            )
        else:
            self.message_user(
                request,
                "No classes created. They may already exist for next year.",
                messages.INFO
            )
    duplicate_class_for_next_year.short_description = "🔄 Duplicate for next year"


@admin.register(Subject)
class SubjectAdmin(admin.ModelAdmin):
    list_display = ('code', 'name', 'department', 'level', 'subject_type')
    list_filter = ('department', 'level', 'subject_type')
    search_fields = ('code', 'name')
    
    # Add these actions to the existing class
    actions = [
        'assign_subjects_to_teachers',
        'view_subject_teachers',
        'view_subject_classes',
        'archive_subjects',
        'delete_selected',
    ]
    
    # 1. Assign Subjects to Teachers
    def assign_subjects_to_teachers(self, request, queryset):
        if request.method == 'POST':
            teacher_id = request.POST.get('teacher_id')
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                for subject in queryset:
                    teacher.subjects.add(subject)
                
                self.message_user(
                    request,
                    f"Successfully assigned {queryset.count()} subject(s) to teacher {teacher.full_name}.",
                    messages.SUCCESS
                )
                return None
            except Teacher.DoesNotExist:
                self.message_user(request, "Teacher not found.", messages.ERROR)
        
        teachers = Teacher.objects.filter(status='ACTIVE')
        return render(
            request,
            'admin/assign_subjects_to_teacher_form.html',
            {
                'subjects': queryset,
                'teachers': teachers,
                'title': 'Assign Subjects to Teacher'
            }
        )
    assign_subjects_to_teachers.short_description = "👨‍🏫 Assign subjects to teacher"
    
    # 2. View Subject Teachers
    def view_subject_teachers(self, request, queryset):
        if queryset.count() == 1:
            subject = queryset.first()
            # Get all teachers who teach this subject
            teachers = Teacher.objects.filter(subjects=subject)
            
            context = {
                'subject': subject,
                'teachers': teachers,
                'title': f'Teachers for {subject.name}'
            }
            return render(request, 'admin/view_subject_teachers.html', context)
        else:
            self.message_user(request, "Please select only ONE subject to view its teachers.", messages.WARNING)
    view_subject_teachers.short_description = "👨‍🏫 View subject teachers"
    
    # 3. View Subject Classes
    def view_subject_classes(self, request, queryset):
        if queryset.count() == 1:
            subject = queryset.first()
            # Get all classes where this subject is taught
            # First, find teachers who teach this subject
            teachers = Teacher.objects.filter(subjects=subject)
            # Then find classes taught by these teachers
            classes = Class.objects.filter(teachers__in=teachers).distinct()
            
            context = {
                'subject': subject,
                'classes': classes,
                'title': f'Classes for {subject.name}'
            }
            return render(request, 'admin/view_subject_classes.html', context)
        else:
            self.message_user(request, "Please select only ONE subject to view its classes.", messages.WARNING)
    view_subject_classes.short_description = "🏫 View subject classes"
    
    # 4. Archive Subjects
    def archive_subjects(self, request, queryset):
        # Note: This requires adding an 'is_active' or 'archived' field to the Subject model
        archived_count = 0
        
        for subject in queryset:
            # In a real implementation, you would have:
            # subject.is_active = False  # or subject.archived = True
            # subject.save()
            archived_count += 1
        
        self.message_user(
            request,
            f"{archived_count} subject(s) marked for archiving. "
            "Note: Add 'is_active' field to Subject model for proper archiving.",
            messages.INFO
        )
    archive_subjects.short_description = "📁 Archive subjects"


@admin.register(Teacher)
class TeacherAdmin(admin.ModelAdmin):
    list_display = ('teacher_id', 'full_name', 'department', 'status', 'employment_date')
    list_filter = ('department', 'status')
    search_fields = ('teacher_id', 'full_name', 'phone_number', 'email')
    filter_horizontal = ('subjects',)
    
    # Added: Make teacher_id read-only since it's auto-generated
    readonly_fields = ('teacher_id',)
    
    # Added: Exclude user field from form (we'll handle it automatically)
    exclude = ('user',)
    
    # Updated actions list for TeacherAdmin
    actions = [
        'bulk_assign_teachers_to_multiple_subjects',
        'generate_teacher_timetable_reports',
        'export_teacher_contact_directory',
        'bulk_update_teacher_employment_status',
        'generate_teacher_performance_reports',
        'assign_teachers_as_class_masters',
        'bulk_send_whatsapp_notifications_to_teachers',
        'delete_selected',
    ]
    
    fieldsets = (
        ('PERSONAL INFORMATION', {
            'fields': ('full_name', 'date_of_birth', 'phone_number', 'email', 'address')
        }),
        ('EMPLOYMENT INFORMATION', {
            'fields': ('teacher_id', 'employment_date', 'department', 'subjects', 'status')
        }),
        ('QUALIFICATIONS & NOTES', {
            'fields': ('qualifications', 'notes'),
            'classes': ('collapse',)
        }),
    )
    
    # 1. Bulk Assign Teachers to Multiple Subjects
    def bulk_assign_teachers_to_multiple_subjects(self, request, queryset):
        if request.method == 'POST':
            subject_ids = request.POST.getlist('subjects')
            if not subject_ids:
                messages.error(request, "Please select at least one subject.")
            else:
                updated_count = 0
                subjects = Subject.objects.filter(id__in=subject_ids)
                
                for teacher in queryset:
                    for subject in subjects:
                        teacher.subjects.add(subject)
                    updated_count += 1
                
                messages.success(
                    request, 
                    f"Successfully assigned {len(subjects)} subjects to {updated_count} teacher(s)."
                )
                return redirect('../')
        
        # Show form with subject selection
        subjects = Subject.objects.all()
        context = {
            'teachers': queryset,
            'subjects': subjects,
            'title': 'Assign Subjects to Teachers'
        }
        return render(request, 'admin/assign_subjects_to_teachers_form.html', context)
    
    bulk_assign_teachers_to_multiple_subjects.short_description = "📚 Bulk assign teachers to subjects"
    
    # 2. Generate Teacher Timetable Reports
    def generate_teacher_timetable_reports(self, request, queryset):
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 750, "Teacher Timetable Reports")
        p.setFont("Helvetica", 10)
        p.drawString(400, 750, f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d')}")
        
        y_position = 700
        
        for teacher in queryset:
            if y_position < 100:
                p.showPage()
                p.setFont("Helvetica", 12)
                y_position = 750
            
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y_position, f"Teacher: {teacher.full_name} ({teacher.teacher_id})")
            y_position -= 20
            
            p.setFont("Helvetica", 12)
            p.drawString(100, y_position, f"Department: {teacher.department}")
            y_position -= 20
            
            # Subjects
            p.drawString(100, y_position, "Subjects:")
            y_position -= 20
            
            subjects = teacher.subjects.all()
            for subject in subjects:
                if y_position < 50:
                    p.showPage()
                    p.setFont("Helvetica", 12)
                    y_position = 750
                
                p.drawString(120, y_position, f"• {subject.code}: {subject.name}")
                y_position -= 15
            
            y_position -= 30
        
        p.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="teacher_timetables.pdf"'
        return response
    
    generate_teacher_timetable_reports.short_description = "📅 Generate teacher timetable reports"
    
    # 3. Export Teacher Contact Directory
    def export_teacher_contact_directory(self, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teachers Directory"
        
        headers = [
            'Teacher ID', 'Full Name', 'Department', 'Phone Number', 
            'Email', 'Employment Date', 'Status', 'Subjects'
        ]
        
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            ws.column_dimensions[col_letter].width = 20
        
        for row_num, teacher in enumerate(queryset, 2):
            ws[f'A{row_num}'] = teacher.teacher_id
            ws[f'B{row_num}'] = teacher.full_name
            ws[f'C{row_num}'] = str(teacher.department) if teacher.department else ''
            ws[f'D{row_num}'] = teacher.phone_number
            ws[f'E{row_num}'] = teacher.email
            ws[f'F{row_num}'] = str(teacher.employment_date) if teacher.employment_date else ''
            ws[f'G{row_num}'] = teacher.status
            ws[f'H{row_num}'] = ', '.join([str(subject) for subject in teacher.subjects.all()])
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="teachers_directory.xlsx"'
        return response
    
    export_teacher_contact_directory.short_description = "📇 Export teacher contact directory"
    
    # 4. Bulk Update Teacher Employment Status
    def bulk_update_teacher_employment_status(self, request, queryset):
        if request.method == 'POST':
            new_status = request.POST.get('new_status')
            updated_count = 0
            
            for teacher in queryset:
                teacher.status = new_status
                teacher.save()
                updated_count += 1
            
            messages.success(request, f"Successfully updated status for {updated_count} teachers to '{new_status}'.")
            return redirect('../')
        
        context = {
            'teachers': queryset,
            'title': 'Bulk Update Teacher Employment Status',
            'status_choices': ['ACTIVE', 'INACTIVE', 'ON_LEAVE', 'TERMINATED', 'RESIGNED']
        }
        return render(request, 'admin/bulk_update_teacher_status_form.html', context)
    
    bulk_update_teacher_employment_status.short_description = "🔄 Bulk update teacher employment status"
    
    # 5. Generate Teacher Performance Reports
    def generate_teacher_performance_reports(self, request, queryset):
        messages.info(
            request,
            f"Performance reports queued for {queryset.count()} teacher(s). "
            "Note: This feature requires Performance model implementation with metrics and evaluations."
        )
    
    generate_teacher_performance_reports.short_description = "📈 Generate teacher performance reports"
    
    # 6. Assign Teachers as Class Masters
    def assign_teachers_as_class_masters(self, request, queryset):
        if request.method == 'POST':
            class_id = request.POST.get('class_id')
            try:
                target_class = Class.objects.get(id=class_id)
                # Add teacher as a teacher to the class (assuming Class model has teachers ManyToManyField)
                for teacher in queryset:
                    target_class.teachers.add(teacher)
                
                messages.success(
                    request, 
                    f"Successfully assigned {queryset.count()} teacher(s) as class masters to {target_class.name}."
                )
                return redirect('../')
            except Class.DoesNotExist:
                messages.error(request, "Selected class does not exist.")
        
        classes = Class.objects.all()
        context = {
            'teachers': queryset,
            'classes': classes,
            'title': 'Assign Teachers as Class Masters'
        }
        return render(request, 'admin/assign_teachers_as_class_masters_form.html', context)
    
    assign_teachers_as_class_masters.short_description = "👨‍🏫 Assign teachers as class masters"
    
    # 7. Bulk Send WhatsApp Notifications to Teachers
    def bulk_send_whatsapp_notifications_to_teachers(self, request, queryset):
        if request.method == 'POST':
            message = request.POST.get('message', '')
            sent_count = 0
            failed_count = 0
            
            for teacher in queryset:
                if teacher.phone_number:
                    try:
                        # WhatsApp integration placeholder
                        phone_number = self._extract_phone_number(teacher.phone_number)
                        print(f"WhatsApp to {teacher.full_name}: {message}")
                        sent_count += 1
                    except Exception as e:
                        print(f"Failed to send to {teacher.full_name}: {str(e)}")
                        failed_count += 1
            
            messages.success(
                request,
                f"Successfully sent WhatsApp notifications to {sent_count} teachers. "
                f"Failed: {failed_count}. Note: Actual WhatsApp integration requires API setup."
            )
            return redirect('../')
        
        context = {
            'teachers': queryset,
            'title': 'Send WhatsApp Notifications to Teachers'
        }
        return render(request, 'admin/send_whatsapp_to_teachers_form.html', context)
    
    bulk_send_whatsapp_notifications_to_teachers.short_description = "📱 Bulk send WhatsApp notifications to teachers"
    
    # Helper method to extract phone number
    def _extract_phone_number(self, contact_string):
        import re
        numbers = re.findall(r'\d+', contact_string)
        if numbers:
            return max(numbers, key=len)
        return None
    
    def save_model(self, request, obj, form, change):
        # If creating a new teacher (not editing an existing one)
        if not change:
            # Check if email is provided
            if obj.email:
                username = obj.email
            else:
                # Generate username from name if email not available
                base_username = obj.full_name.replace(' ', '_').lower()
                username = base_username
                counter = 1
                # Ensure unique username
                while User.objects.filter(username=username).exists():
                    username = f"{base_username}_{counter}"
                    counter += 1
            
            # Generate random password using Django's built-in get_random_string
            password = get_random_string(length=10)
            
            # Create User object
            user = User.objects.create_user(
                username=username,
                email=obj.email if obj.email else '',
                password=password,
                first_name=obj.full_name.split()[0] if obj.full_name and ' ' in obj.full_name else obj.full_name or '',
                last_name=' '.join(obj.full_name.split()[1:]) if obj.full_name and ' ' in obj.full_name else ''
            )
            
            # Assign the created user to the teacher
            obj.user = user
        
        # Save the teacher object
        super().save_model(request, obj, form, change)


# Unregister the default User admin
admin.site.unregister(User)

# Register custom User admin with new actions
@admin.register(User)
class CustomUserAdmin(BaseUserAdmin):
    actions = [
        'enable_users',
        'disable_users', 
        'reset_passwords',
        'send_login_credentials',
        'assign_to_group',
        'remove_from_group',
        'mark_as_staff',
        'remove_staff_status',
        'delete_selected',
    ]
    
    def enable_users(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f'{updated} user(s) enabled successfully.')
    enable_users.short_description = "✅ Enable selected users"
    
    def disable_users(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} user(s) disabled successfully.')
    disable_users.short_description = "🚫 Disable selected users"
    
    def reset_passwords(self, request, queryset):
        for user in queryset:
            new_password = get_random_string(length=10)
            user.set_password(new_password)
            user.save()
        self.message_user(request, f'Passwords reset for {queryset.count()} user(s).')
    reset_passwords.short_description = "🔑 Reset passwords"
    
    def send_login_credentials(self, request, queryset):
        for user in queryset:
            # In a real application, you would send an email here
            print(f"Login credentials sent to {user.email}: Username: {user.username}")
        self.message_user(request, f'Login credentials sent to {queryset.count()} user(s).')
    send_login_credentials.short_description = "📧 Send login credentials"
    
    def assign_to_group(self, request, queryset):
        if request.method == 'POST':
            group_id = request.POST.get('group_id')
            try:
                group = Group.objects.get(id=group_id)
                for user in queryset:
                    user.groups.add(group)
                self.message_user(request, f'Added {queryset.count()} user(s) to group "{group.name}".')
                return None
            except Group.DoesNotExist:
                self.message_user(request, "Group not found.", messages.ERROR)
        
        groups = Group.objects.all()
        return render(
            request,
            'admin/assign_to_group_form.html',
            {
                'users': queryset,
                'groups': groups,
                'title': 'Assign Users to Group'
            }
        )
    assign_to_group.short_description = "👥 Assign to group"
    
    def remove_from_group(self, request, queryset):
        if request.method == 'POST':
            group_id = request.POST.get('group_id')
            try:
                group = Group.objects.get(id=group_id)
                for user in queryset:
                    user.groups.remove(group)
                self.message_user(request, f'Removed {queryset.count()} user(s) from group "{group.name}".')
                return None
            except Group.DoesNotExist:
                self.message_user(request, "Group not found.", messages.ERROR)
        
        groups = Group.objects.all()
        return render(
            request,
            'admin/remove_from_group_form.html',
            {
                'users': queryset,
                'groups': groups,
                'title': 'Remove Users from Group'
            }
        )
    remove_from_group.short_description = "👤 Remove from group"
    
    def mark_as_staff(self, request, queryset):
        updated = queryset.update(is_staff=True)
        self.message_user(request, f'{updated} user(s) granted staff access.')
    mark_as_staff.short_description = "👨‍💼 Mark as staff"
    
    def remove_staff_status(self, request, queryset):
        updated = queryset.update(is_staff=False)
        self.message_user(request, f'{updated} user(s) removed from staff.')
    remove_staff_status.short_description = "📊 Remove staff status"